from openpyxl import load_workbook

input_file  = r"c:\dev\masuot\simulator-mi.xlsx"
output_file = r"c:\dev\masuot\simulator-mi-anonymized.xlsx"

wb = load_workbook(input_file)
ws = wb["תקציבים"]

for row in range(5,252):
    code = ws[f"A{row}"].value
    ws[f"B{row}"].value = "משפחה" + "_" + str(code)

ws = wb["ריכוז נתונים"]
for row in range(5,164):
    code = ws[f"A{row}"].value
    ws[f"B{row}"].value = "משפחה" + "_" + str(code)

ws = wb["חברים"]
counter_codes = {}
for row in range(6,711):
    code = ws[f"H{row}"].value #קוד תקציב
    id_value  = ws[f"L{row}"].value
    sex  = ws[f"E{row}"].value
    family = ws[f"N{row}"].value
    status = ws[f"Q{row}"].value

    #id
    ws[f"L{row}"].value = str(id_value)[:2] + "XXXXX" + str(id_value)[-2:]
    #name
    ws[f"B{row}"].value = 'XXXXX'
    #family name
    ws[f"C{row}"].value = 'XXXXX'
    #first name
    ws[f"D{row}"].value = 'XXXXX'
    #family
    ws[f"N{row}"].value = 'XXXXX'
    #desc
    if code not in counter_codes:
        counter_codes[code] = 1
    else:
        counter_codes[code] += 1
    ws[f"I{row}"].value = "משפחה" + "_" + str(code) + "_" + str(counter_codes[code])

ws = wb["זיכוי תקציב מפורט"]
counter_codes = {} 
for row in range(5,113765):
    code = ws[f"A{row}"].value 
    if code not in counter_codes:
        counter_codes[code] = 1
    else:
        counter_codes[code] += 1
    ws[f"B{row}"].value = "משפחה" + "_" + str(code) + "_" + str(counter_codes[code]) 

ws = wb["הכנסות שכר "]
counter_codes = {} 
for row in range(9,304):
    code = ws[f"A{row}"].value 
    if code not in counter_codes:
        counter_codes[code] = 1
    else:
        counter_codes[code] += 1
    ws[f"B{row}"].value = "משפחה" + "_" + str(counter_codes[code])

ws = wb["נתוני עבודה 25"]
counter_codes = {}
for row in range(3,298):
    code = ws[f"B{row}"].value
    if code not in counter_codes:
        counter_codes[code] = 1
    else:
        counter_codes[code] += 1 
    ws[f"C{row}"].value = "משפחה" + "_" + str(code)
    ws[f"D{row}"].value = "משפחה" + "_" + str(code) + "_" + str(counter_codes[code])

wb.save(output_file)
print("נוצר קובץ חדש:")
print(output_file)