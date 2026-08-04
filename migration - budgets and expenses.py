import psycopg2
from psycopg2.extras import execute_values
from openpyxl import load_workbook

EXCEL_PATH = "masuot-data.xlsx"
#EXCEL_PATH = r"c:\users\hagai\Downloads\פרק א - תקציבים והוצאות גרסה 27 להצגה.xlsx"

def connect():
    return psycopg2.connect(
        host="aws-1-ap-northeast-1.pooler.supabase.com",
        port=5432,
        database="postgres",
        user="postgres.jhkxyiiwtxtgqxovljkl",
        password="__Por@t2019!"
    )

def to_num(v):
    try:
        return float(v)
    except:
        return 0

def is_valid(v):
    return v is not None and str(v).strip() != ""

def norm_id(v):
    if v is None:
        return None

    try:
        return str(int(float(v))).strip()
    except:
        return str(v).strip()

def ensure_column_exists(cur):

    cur.execute("""
        SELECT column_name
        FROM information_schema.columns
        WHERE table_name='families'
        AND column_name='health_participation'
    """)

    if not cur.fetchone():

        print("Adding column health_participation...")

        cur.execute("""
            ALTER TABLE families
            ADD COLUMN health_participation NUMERIC DEFAULT 0
        """)

    cur.execute("""
        SELECT column_name
        FROM information_schema.columns
        WHERE table_name='families'
        AND column_name='mutual_responsibility_cap'
    """)

    if not cur.fetchone():

        print("Adding column mutual_responsibility_cap...")

        cur.execute("""
            ALTER TABLE families
            ADD COLUMN mutual_responsibility_cap NUMERIC DEFAULT 0
        """)

def main():

    conn = connect()
    cur = conn.cursor()

    wb = load_workbook(
        EXCEL_PATH,
        data_only=True
    )

    summary = wb["ריכוז נתונים"]

    members_sheet = wb["חברים"]

    salary_sheet = wb["הכנסות שכר "]

    discount_sheet = wb["הנחות "]

    ensure_column_exists(cur)

    print("Reset DB...")

    # =====================================
    # DROP
    # =====================================

    cur.execute("DROP TABLE IF EXISTS members CASCADE")
    cur.execute("DROP TABLE IF EXISTS families CASCADE")
    cur.execute("DROP TABLE IF EXISTS rules CASCADE")

    print("Create tables...")

    # =====================================
    # FAMILIES
    # =====================================

    cur.execute("""
        CREATE TABLE families (

            budget_code TEXT PRIMARY KEY,

            family_name TEXT,

            family_standard NUMERIC,
            income_for_standard NUMERIC,
            budget_distribution NUMERIC,
            personal_bonus NUMERIC,
            women_work_benefit NUMERIC,
            travel NUMERIC,
            periodic_grant NUMERIC,
            special_help NUMERIC,
            current_state NUMERIC,

            pension NUMERIC,
            survivors NUMERIC,
            old_age_allowance NUMERIC,
            child_allowance NUMERIC,

            community_tax NUMERIC,
            municipal_tax NUMERIC,
            arnona NUMERIC,

            health_total NUMERIC,
            health_0_50 NUMERIC,
            health_50_70 NUMERIC,
            health_70_plus NUMERIC,

            toddlers NUMERIC,
            kindergarten NUMERIC,
            elementary NUMERIC,
            middle NUMERIC,
            high NUMERIC,

            health_participation NUMERIC,
            mutual_responsibility_cap NUMERIC,

            hishtalmut_fund NUMERIC,
            pension_contribution NUMERIC,

            additional_allowances NUMERIC
        )
        """)

    # =====================================
    # MEMBERS
    # =====================================

    cur.execute("""
    CREATE TABLE members (

        id SERIAL PRIMARY KEY,

        budget_code TEXT,

        member_code TEXT,

        first_name TEXT,

        last_name TEXT,

        age NUMERIC,

        net_salary NUMERIC,

        gross_income NUMERIC,

        credit_points NUMERIC,

        income_type TEXT,

        status_code NUMERIC,

        education_group TEXT
    )
    """)

    cur.execute("""
    ALTER TABLE members
    ADD CONSTRAINT uq_members_budget_member
    UNIQUE (budget_code, member_code)
    """)

    cur.execute("""
    CREATE INDEX idx_members_budget_code
    ON members(budget_code)
    """)

    cur.execute("""
    CREATE INDEX idx_members_member_code
    ON members(member_code)
    """)

    # =====================================
    # RULES
    # =====================================

    cur.execute("""
    CREATE TABLE rules (

        key TEXT PRIMARY KEY,

        value NUMERIC,

        category TEXT
    )
    """)

    # =====================================
    # IMPORT RULES
    # =====================================

    print("Importing rules...")

    rules_data = []

    # =====================================
    # EDUCATION
    # =====================================

    row = next(summary.iter_rows(min_row=1))

    rules_data.extend([

        ("nursery", to_num(row[7].value), "education"),

        ("kindergarten", to_num(row[8].value), "education"),

        ("primary", to_num(row[9].value), "education"),

        ("middle", to_num(row[10].value), "education"),

        ("highschool", to_num(row[11].value), "education")
    ])

    # =====================================
    # HEALTH
    # =====================================

    rules_data.extend([

        ("health_total", to_num(row[38].value), "health"),
        ("health_0_50", to_num(row[39].value), "health"),
        ("health_50_70", to_num(row[40].value), "health"),
        ("health_70_plus", to_num(row[41].value), "health")
    ])

    # =====================================
    # DISCOUNTS
    # =====================================

    rules_data.extend([

        ("K5", to_num(discount_sheet.cell(row=5, column=11).value), "discount"),

        ("L5", to_num(discount_sheet.cell(row=5, column=12).value), "discount"),

        ("K6", to_num(discount_sheet.cell(row=6, column=11).value), "discount"),

        ("J6", to_num(discount_sheet.cell(row=6, column=10).value), "discount"),

        ("L6", to_num(discount_sheet.cell(row=6, column=12).value), "discount"),

        ("M5", to_num(discount_sheet.cell(row=5, column=13).value), "discount"),

        ("K7", to_num(discount_sheet.cell(row=7, column=11).value), "discount"),

        ("J7", to_num(discount_sheet.cell(row=7, column=10).value), "discount"),

        ("L7", to_num(discount_sheet.cell(row=7, column=12).value), "discount"),

        ("M6", to_num(discount_sheet.cell(row=6, column=13).value), "discount"),

        ("F16", to_num(discount_sheet.cell(row=16, column=6).value), "discount"),

        ("F21", to_num(discount_sheet.cell(row=21, column=6).value), "discount"),

        ("F19", to_num(discount_sheet.cell(row=19, column=6).value), "discount"),

        ("F4", to_num(discount_sheet.cell(row=4, column=6).value), "discount")
    ])

    # =====================================
    # TAX BRACKETS (S:U)
    # =====================================

    for idx, row in enumerate(

        salary_sheet.iter_rows(
            min_row=2,
            max_row=7,
            min_col=19,
            max_col=21
        ),

        start=1
    ):

        threshold = to_num(row[0].value)

        tax_rate = to_num(row[1].value)

        diff = to_num(row[2].value)

        rules_data.extend([

            (
                f"mas_{idx}_limit",
                threshold,
                "tax"
            ),

            (
                f"mas_{idx}_rate",
                tax_rate,
                "tax"
            ),

            (
                f"mas_{idx}_diff",
                diff,
                "tax"
            )
        ])

    # =====================================
    # PAYROLL PARAMETERS (W:X)
    # =====================================

    for row in salary_sheet.iter_rows(
        min_row=2,
        max_row=7,
        min_col=23,
        max_col=24
    ):

        key = row[0].value

        value = to_num(row[1].value)

        if not is_valid(key):
            continue

        rules_data.append((
            str(key).strip(),
            value,
            "payroll"
        ))

    # =====================================
    # SAVINGS PARAMETERS (P:Q)
    # =====================================

    hishtalmut_rate = to_num(
        salary_sheet["P2"].value
    )

    hishtalmut_min_salary = to_num(
        salary_sheet["Q2"].value
    )

    pension_fund_rate = to_num(
        salary_sheet["P3"].value
    )

    pension_fund_offset = to_num(
        salary_sheet["Q3"].value
    )

    rules_data.extend([

        (
            "hishtalmut_rate",
            hishtalmut_rate,
            "savings"
        ),

        (
            "hishtalmut_min_salary",
            hishtalmut_min_salary,
            "savings"
        ),

        (
            "pension_fund_rate",
            pension_fund_rate,
            "savings"
        ),

        (
            "pension_fund_offset",
            pension_fund_offset,
            "savings"
        )
    ])

    # =====================================
    # INSERT RULES
    # =====================================

    print("Rules:", len(rules_data))

    execute_values(cur, """
        INSERT INTO rules (

            key,

            value,

            category

        ) VALUES %s
    """, rules_data)

  

    # =====================================
    # SAVINGS MAP
    # =====================================

    savings_map = {}

    for row in salary_sheet.iter_rows(min_row=9):

        code = row[0].value

        if not is_valid(code):
            continue

        code = str(code).strip()

        hishtalmut = to_num(row[14].value)

        pension = to_num(row[15].value)

        if code not in savings_map:

            savings_map[code] = {
                "hishtalmut": 0,
                "pension": 0
            }

        savings_map[code]["hishtalmut"] += hishtalmut

        savings_map[code]["pension"] += pension

    # =====================================
    # FAMILIES
    # =====================================
    print("Importing families...")

    families_data = []

    for row in summary.iter_rows(min_row=5):

        code = row[0].value
        name = row[1].value

        if not is_valid(code):
            continue

        if name and "סיכום" in str(name):
            continue

        code = str(code).strip()

        health_participation = to_num(row[43].value)
        mutual_responsibility_cap = to_num(row[45].value)

        families_data.append((

            code,
            name,

            to_num(row[14].value),
            to_num(row[30].value),
            to_num(row[15].value),
            to_num(row[18].value),
            to_num(row[19].value),
            to_num(row[20].value),
            to_num(row[17].value),
            to_num(row[16].value),
            to_num(row[21].value),

            to_num(row[24].value),
            to_num(row[25].value),
            to_num(row[26].value),

            to_num(row[28].value),  # child_allowance (AC)
            to_num(row[27].value),  # additional_allowances (AB)

            to_num(row[46].value),
            to_num(row[47].value),
            to_num(row[48].value),

            to_num(row[38].value),
            to_num(row[39].value),
            to_num(row[40].value),
            to_num(row[41].value),

            to_num(row[7].value),
            to_num(row[8].value),
            to_num(row[9].value),
            to_num(row[10].value),
            to_num(row[11].value),

            health_participation,
            mutual_responsibility_cap,

            savings_map.get(code, {}).get("hishtalmut", 0),
            savings_map.get(code, {}).get("pension", 0)
        ))

    execute_values(cur, """
    INSERT INTO families (

        budget_code,
        family_name,
        family_standard,
        income_for_standard,
        budget_distribution,
        personal_bonus,
        women_work_benefit,
        travel,
        periodic_grant,
        special_help,
        current_state,
        pension,
        survivors,
        old_age_allowance,
        child_allowance,
        additional_allowances,
        community_tax,
        municipal_tax,
        arnona,
        health_total,
        health_0_50,
        health_50_70,
        health_70_plus,
        toddlers,
        kindergarten,
        elementary,
        middle,
        high,
        health_participation,
        mutual_responsibility_cap,
        hishtalmut_fund,
        pension_contribution

    ) VALUES %s
    """, families_data)

    print("Families:", len(families_data))

    # =====================================
    # MEMBERS
    # =====================================

    print("Importing members...")

    salary_map = {}

    for row in salary_sheet.iter_rows(min_row=9):

        code = row[0].value

        member_code = norm_id(row[2].value)

        if not is_valid(code) or not is_valid(member_code):
            continue

        code = str(code).strip()

        key = (code, member_code)

        gross_income = to_num(row[4].value)      # E

        income_type = row[8].value               # I

        credit_points = to_num(row[10].value)   # K

        net_income = to_num(row[12].value)      # M

        salary_map[key] = {

            "salary": salary_map.get(key, {}).get("salary", 0) + net_income,

            "gross_income": gross_income,

            "net_income": net_income,

            "credit_points": credit_points,

            "income_type": income_type
        }

    members_data = []
    seen = set()

    for row in members_sheet.iter_rows(min_row=6):

        code = row[7].value
        member_code = norm_id(row[0].value)

        first_name = row[1].value
        last_name = row[2].value
        age = to_num(row[6].value)
        status_code = to_num(row[15].value)
        education_group = row[20].value

        if not is_valid(code) or not is_valid(member_code):
            continue

        code = str(code).strip()

        key = (code, member_code)

        if key in seen:
            continue

        seen.add(key)

        salary_info = salary_map.get(key, {})

        net_salary = salary_info.get("net_income", 0)
        gross_income = salary_info.get("gross_income", 0)
        credit_points = salary_info.get("credit_points", 0)
        income_type = salary_info.get("income_type", "")

        members_data.append((

            code,
            member_code,
            first_name,
            last_name,
            age,
            net_salary,
            gross_income,
            credit_points,
            income_type,
            status_code,
            education_group
        ))

    execute_values(cur, """
        INSERT INTO members (

            budget_code,

            member_code,

            first_name,

            last_name,

            age,

            net_salary,

            gross_income,

            credit_points,

            income_type,

            status_code,

            education_group

        ) VALUES %s
    """, members_data)

    print("Members:", len(members_data))

    conn.commit()

    conn.close()

    print("✅ DONE")

if __name__ == "__main__":
    main()