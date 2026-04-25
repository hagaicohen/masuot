import psycopg2
from psycopg2.extras import execute_values
from openpyxl import load_workbook
import bcrypt
from datetime import datetime
from concurrent.futures import ThreadPoolExecutor
from datetime import datetime, timezone

EXCEL_PATH = "passwords.xlsx"

def connect():
    return psycopg2.connect(
        host="aws-1-ap-northeast-1.pooler.supabase.com",
        port=5432,
        database="postgres",
        user="postgres.jhkxyiiwtxtgqxovljkl",
        password="__Por@t2019!"
    )
BATCH_SIZE = 100  # 🔥 יציב ל-Supabase

def to_str(v):
    if v is None:
        return ""

    s = str(v)

    # 🔥 ניקוי אגרסיבי
    s = s.replace('\n', '').replace('\r', '')
    s = s.replace('\u200e', '').replace('\u200f', '')  # RTL marks

    return s.strip()

def hash_password(password: str) -> str:
    return bcrypt.hashpw(
        password.encode('utf-8'),
        bcrypt.gensalt(10)  # 🔥 מהיר מספיק ובטוח
    ).decode('utf-8')

def main():
    conn = connect()
    cur = conn.cursor()

    # 🔥 DROP
    cur.execute("DROP TABLE IF EXISTS users CASCADE")

    # 🔥 CREATE
    cur.execute("""
    CREATE TABLE users (
        id SERIAL PRIMARY KEY,
        budget_code TEXT UNIQUE,
        description TEXT,
        password_hash TEXT,
        role TEXT,
        last_login TIMESTAMP,
        created_at TIMESTAMP,
        updated_at TIMESTAMP
    )
    """)

    # 🔥 INDEXES
    cur.execute("CREATE UNIQUE INDEX idx_users_budget_code ON users(budget_code)")
    cur.execute("CREATE INDEX idx_users_role ON users(role)")
    cur.execute("CREATE INDEX idx_users_last_login ON users(last_login)")

    wb = load_workbook(EXCEL_PATH, data_only=True)
    sheet = wb["סיסמאות"]

    rows = list(sheet.iter_rows(min_row=3, values_only=True))
    print(f"Rows: {len(rows)}")

    now = datetime.now(timezone.utc)

    users_data = []

    # =========================
    # BUILD DATA
    # =========================
    for i, r in enumerate(rows, start=1):
        user_name = to_str(r[0])     # USER_NAME → budget_code
        description = to_str(r[1])   # TEUR
        password_raw = to_str(r[2])  # PASSWORD

        if not user_name or not password_raw:
            continue

        if i % 100 == 0:
            print(f"Processing {i}/{len(rows)}")

        password_hash = hash_password(password_raw)

        users_data.append((
            user_name,
            description,
            password_hash,
            'member',
            now,
            now
        ))

    print(f"Users ready: {len(users_data)}")

    if not users_data:
        print("🚨 No users to insert")
        return

    # =========================
    # BATCH PROCESS
    # =========================
    for i in range(0, len(users_data), BATCH_SIZE):
        batch = users_data[i:i + BATCH_SIZE]

        print(f"🚀 Batch {i}-{i + len(batch)}")

        # -------------------------
        # INSERT חדשים בלבד
        # -------------------------
        execute_values(cur, """
            INSERT INTO users (
                budget_code,
                description,
                password_hash,
                role,
                created_at,
                updated_at
            ) VALUES %s
            ON CONFLICT (budget_code) DO NOTHING
        """, batch)

        # -------------------------
        # UPDATE קיימים
        # -------------------------
        for u in batch:
            cur.execute("""
                UPDATE users
                SET
                    description = %s,
                    password_hash = %s,
                    updated_at = %s
                WHERE budget_code = %s
            """, (u[1], u[2], u[4], u[0]))

        conn.commit()  # 🔥 חשוב — משחרר עומס

    conn.close()

    print("✅ USERS MIGRATION DONE")

if __name__ == "__main__":
    main()