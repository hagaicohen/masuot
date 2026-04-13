import psycopg2
from psycopg2.extras import execute_values
from openpyxl import load_workbook

EXCEL_PATH = "masuot-data.xlsx"

def connect():
    return psycopg2.connect(
        host="aws-1-ap-northeast-1.pooler.supabase.com",
        port=5432,
        database="postgres",
        user="postgres.jhkxyiiwtxtgqxovljkl",
        password="__Por@t2019!")


def to_num(v):
    try:
        return float(v)
    except:
        return 0

def is_valid(v):
    return v is not None and str(v).strip() != ""

def norm_id(v):
    if v is None:
        return None
    try:
        return str(int(float(v))).strip()
    except:
        return str(v).strip()

def ensure_column_exists(cur):
    """
    🔥 UPDATED — now ensures BOTH columns exist
    """

    # 🔥 EXISTING
    cur.execute("""
        SELECT column_name 
        FROM information_schema.columns 
        WHERE table_name='families' AND column_name='health_participation'
    """)
    if not cur.fetchone():
        print("Adding column health_participation...")
        cur.execute("""
            ALTER TABLE families 
            ADD COLUMN health_participation NUMERIC DEFAULT 0
        """)

    # 🔥 ADDED — AS column (mutual responsibility cap)
    cur.execute("""
        SELECT column_name 
        FROM information_schema.columns 
        WHERE table_name='families' AND column_name='mutual_responsibility_cap'
    """)
    if not cur.fetchone():
        print("Adding column mutual_responsibility_cap...")
        cur.execute("""
            ALTER TABLE families 
            ADD COLUMN mutual_responsibility_cap NUMERIC DEFAULT 0
        """)

def main():
    conn = connect()
    cur = conn.cursor()

    wb = load_workbook(EXCEL_PATH, data_only=True)

    summary = wb["ריכוז נתונים"]
    members_sheet = wb["חברים"]
    salary_sheet = wb["הכנסות שכר "]
    discount_sheet = wb["הנחות "]

    ensure_column_exists(cur)

    print("Reset DB...")

    cur.execute("DROP TABLE IF EXISTS salary_income CASCADE")
    cur.execute("DROP TABLE IF EXISTS members CASCADE")
    cur.execute("DROP TABLE IF EXISTS families CASCADE")
    cur.execute("DROP TABLE IF EXISTS rules CASCADE")

    print("Create tables...")

    cur.execute("""
    CREATE TABLE families (
        budget_code TEXT PRIMARY KEY,
        family_name TEXT,
        family_standard NUMERIC,
        income_for_standard NUMERIC,
        budget_distribution NUMERIC,
        personal_bonus NUMERIC,
        women_work_benefit NUMERIC,
        travel NUMERIC,
        periodic_grant NUMERIC,
        special_help NUMERIC,
        current_state NUMERIC,
        pension NUMERIC,
        survivors NUMERIC,
        old_age_allowance NUMERIC,
        child_allowance NUMERIC,
        community_tax NUMERIC,
        municipal_tax NUMERIC,
        arnona NUMERIC,
        health_total NUMERIC,
        health_0_50 NUMERIC,
        health_50_70 NUMERIC,
        health_70_plus NUMERIC,
        toddlers NUMERIC,
        kindergarten NUMERIC,
        elementary NUMERIC,
        middle NUMERIC,
        high NUMERIC,
        health_participation NUMERIC,
        mutual_responsibility_cap NUMERIC,
        hishtalmut_fund NUMERIC,
        pension_contribution NUMERIC
    )
    """)

    cur.execute("""
    CREATE TABLE members (
        id SERIAL PRIMARY KEY,
        budget_code TEXT,
        member_code TEXT,
        first_name TEXT,
        last_name TEXT,
        age NUMERIC,
        net_salary NUMERIC,
        status_code NUMERIC,
        education_group TEXT        
    )
    """)

    cur.execute("""
    CREATE TABLE salary_income (
        id SERIAL PRIMARY KEY,
        budget_code TEXT,
        member_code TEXT,
        amount NUMERIC
    )
    """)

    # MEMBERS
    cur.execute("CREATE INDEX idx_members_budget_code ON members(budget_code)")
    cur.execute("CREATE INDEX idx_members_member_code ON members(member_code)")

    # SALARY
    cur.execute("CREATE INDEX idx_salary_budget_code ON salary_income(budget_code)")
    cur.execute("CREATE INDEX idx_salary_member_code ON salary_income(member_code)")

    # 🔥 חשוב לביצועים (JOIN)
    cur.execute("""
    CREATE INDEX idx_salary_budget_member 
    ON salary_income(budget_code, member_code)
    """)

    savings_map = {}
    for row in salary_sheet.iter_rows(min_row=9):
        code = row[0].value

        if not is_valid(code):
            continue

        code = str(code).strip()

        hishtalmut = to_num(row[14].value)
        pension = to_num(row[15].value)

        if code not in savings_map:
            savings_map[code] = {"hishtalmut": 0, "pension": 0}

        savings_map[code]["hishtalmut"] += hishtalmut
        savings_map[code]["pension"] += pension

    families_data = []

    for row in summary.iter_rows(min_row=5):
        code = row[0].value
        name = row[1].value

        if not is_valid(code):
            continue
        if name and "סיכום" in str(name):
            continue

        code = str(code).strip()

        health_participation = to_num(row[42].value)
        mutual_responsibility_cap = to_num(row[44].value)

        families_data.append((
            code,
            name,
            to_num(row[14].value),
            to_num(row[29].value),

            to_num(row[15].value),
            to_num(row[18].value),
            to_num(row[19].value),
            to_num(row[20].value),
            to_num(row[17].value),
            to_num(row[16].value),
            to_num(row[21].value),
            to_num(row[24].value),
            to_num(row[25].value),
            to_num(row[26].value),
            to_num(row[27].value),

            to_num(row[45].value),
            to_num(row[46].value),
            to_num(row[47].value),

            to_num(row[37].value),
            to_num(row[38].value),
            to_num(row[39].value),
            to_num(row[40].value),

            to_num(row[7].value),
            to_num(row[8].value),
            to_num(row[9].value),
            to_num(row[10].value),
            to_num(row[11].value),

            health_participation,
            mutual_responsibility_cap,

            savings_map.get(code, {}).get("hishtalmut", 0),
            savings_map.get(code, {}).get("pension", 0)
        ))

    execute_values(cur, """
        INSERT INTO families (
            budget_code,
            family_name,
            family_standard,
            income_for_standard,
            budget_distribution,
            personal_bonus,
            women_work_benefit,
            travel,
            periodic_grant,
            special_help,
            current_state,
            pension,
            survivors,
            old_age_allowance,
            child_allowance,
            community_tax,
            municipal_tax,
            arnona,

            health_total,
            health_0_50,
            health_50_70,
            health_70_plus,

            toddlers,
            kindergarten,
            elementary,
            middle,
            high,

            health_participation,
            mutual_responsibility_cap,

            hishtalmut_fund,
            pension_contribution
        ) VALUES %s
    """, families_data)

    print("Families:", len(families_data))

    # =========================
    # MEMBERS (unchanged)
    # =========================
    print("Importing members...")

    salary_map = {}

    for row in salary_sheet.iter_rows(min_row=9):
        code = row[0].value
        member_code = norm_id(row[2].value)

        if not is_valid(code) or not is_valid(member_code):
            continue

        code = str(code).strip()

        key = (code, member_code)
        salary = to_num(row[13].value)

        salary_map[key] = salary_map.get(key, 0) + salary

    members_data = []

    for row in members_sheet.iter_rows(min_row=6):
        code = row[7].value
        member_code = norm_id(row[0].value)
        first_name = row[1].value
        last_name = row[2].value
        age = to_num(row[6].value)
        status_code = to_num(row[15].value)
        education_group = row[20].value

        if not is_valid(code) or not is_valid(member_code):
            continue

        code = str(code).strip()

        key = (code, member_code)
        net_salary = salary_map.get(key, 0)

        members_data.append((
            code,
            member_code,
            first_name,
            last_name,
            age,
            net_salary,
            status_code,
            education_group
        ))

    execute_values(cur, """
        INSERT INTO members (
            budget_code,
            member_code,
            first_name,
            last_name,
            age,
            net_salary,
            status_code,
            education_group
        ) VALUES %s
    """, members_data)

    print("Members:", len(members_data))

    # =========================
    # 🔥 ADDED — RULES (בלבד)
    # =========================
    print("Rebuilding rules...")

    cur.execute("DROP TABLE IF EXISTS rules")

    cur.execute("""
    CREATE TABLE rules (
        key TEXT PRIMARY KEY,
        nursery NUMERIC,
        kindergarten NUMERIC,
        "primary" NUMERIC,
        middle NUMERIC,
        highschool NUMERIC,
        health_total NUMERIC,
        health_0_50 NUMERIC,
        health_50_70 NUMERIC,
        health_70_plus NUMERIC,
        K5 NUMERIC, L5 NUMERIC,
        K6 NUMERIC, J6 NUMERIC, L6 NUMERIC, M5 NUMERIC,
        K7 NUMERIC, J7 NUMERIC, L7 NUMERIC, M6 NUMERIC,
        F16 NUMERIC, F21 NUMERIC,F19 NUMERIC, F4 NUMERIC

    )
    """)

    row = next(summary.iter_rows(min_row=1))

    cur.execute("""
    INSERT INTO rules VALUES (%s, %s, %s, %s, %s, %s,
                             %s, %s, %s, %s,
                             %s, %s, %s, %s, %s, %s,
                             %s, %s, %s, %s,
                             %s, %s,%s,%s
                             )
    """, (
        "default",
        to_num(row[7].value),
        to_num(row[8].value),
        to_num(row[9].value),
        to_num(row[10].value),
        to_num(row[11].value),

        to_num(row[37].value),
        to_num(row[38].value),
        to_num(row[39].value),
        to_num(row[40].value),

        to_num(discount_sheet.cell(row=5, column=11).value),
        to_num(discount_sheet.cell(row=5, column=12).value),
        to_num(discount_sheet.cell(row=6, column=11).value),
        to_num(discount_sheet.cell(row=6, column=10).value),
        to_num(discount_sheet.cell(row=6, column=12).value),
        to_num(discount_sheet.cell(row=5, column=13).value),

        to_num(discount_sheet.cell(row=7, column=11).value),
        to_num(discount_sheet.cell(row=7, column=10).value),
        to_num(discount_sheet.cell(row=7, column=12).value),
        to_num(discount_sheet.cell(row=6, column=13).value),

        to_num(discount_sheet.cell(row=16, column=6).value),
        to_num(discount_sheet.cell(row=21, column=6).value),
        to_num(discount_sheet.cell(row=19, column=6).value),
        to_num(discount_sheet.cell(row=4, column=6).value),
    ))

    conn.commit()
    conn.close()

    print("✅ DONE")

if __name__ == "__main__":
    main()