import psycopg2
from psycopg2.extras import execute_values
from openpyxl import load_workbook

EXCEL_PATH = "masuot-special-budget.xlsx"

def connect():
    return psycopg2.connect(
        host="aws-1-ap-northeast-1.pooler.supabase.com",
        port=5432,
        database="postgres",
        user="postgres.jhkxyiiwtxtgqxovljkl",
        password="__Por@t2019!")

def to_num(v):
    try:
        return float(v)
    except:
        return None

def is_valid(v):
    return v is not None and str(v).strip() != ""

def norm_id(v):
    if v is None:
        return None
    try:
        return str(int(float(v))).strip()
    except:
        return str(v).strip()

def main():
    conn = connect()
    cur = conn.cursor()

    wb = load_workbook(EXCEL_PATH, data_only=True)
    members_sheet = wb["חברים"]

    print("Reset special budget table...")

    cur.execute("DROP TABLE IF EXISTS special_budgets CASCADE")

    print("Create table...")

    cur.execute("""
    CREATE TABLE special_budgets (
        budget_code TEXT,
        member_code TEXT,
        last_name TEXT,
        first_name TEXT,
        birth_date DATE,
        age NUMERIC,

        bar_mitzvah_amount NUMERIC,
        bar_mitzvah_year NUMERIC,

        bat_mitzvah_amount NUMERIC,
        bat_mitzvah_year NUMERIC,

        wedding_grant NUMERIC,
        wedding_year NUMERIC,

        study_grant NUMERIC,
        study_year NUMERIC,

        PRIMARY KEY (budget_code, member_code)
    )
    """)

    # INDEXES
    cur.execute("CREATE INDEX idx_sb_budget_code ON special_budgets(budget_code)")
    cur.execute("CREATE INDEX idx_sb_member_code ON special_budgets(member_code)")
    cur.execute("""
        CREATE INDEX idx_sb_budget_member 
        ON special_budgets(budget_code, member_code)
    """)

    print("Importing special budgets...")

    data = []

    for row in members_sheet.iter_rows(min_row=6):
        budget_code = row[19].value  # 🔥 עמודה T
        member_code = norm_id(row[0].value)

        if not is_valid(budget_code) or not is_valid(member_code):
            continue

        budget_code = str(budget_code).strip()

        data.append((
            budget_code,
            member_code,
            row[2].value,  # last_name
            row[3].value,  # first_name
            row[5].value,  # birth_date
            to_num(row[6].value),  # age

            to_num(row[27].value),  # AB
            to_num(row[28].value),  # AC

            to_num(row[29].value),  # AD
            to_num(row[30].value),  # AE

            to_num(row[31].value),  # AF
            to_num(row[32].value),  # AG

            to_num(row[33].value),  # AH
            to_num(row[34].value),  # AI
        ))

    execute_values(cur, """
        INSERT INTO special_budgets (
            budget_code,
            member_code,
            last_name,
            first_name,
            birth_date,
            age,
            bar_mitzvah_amount,
            bar_mitzvah_year,
            bat_mitzvah_amount,
            bat_mitzvah_year,
            wedding_grant,
            wedding_year,
            study_grant,
            study_year
        ) VALUES %s
    """, data)

    print("Inserted:", len(data))

    conn.commit()
    conn.close()

    print("✅ DONE")

if __name__ == "__main__":
    main()