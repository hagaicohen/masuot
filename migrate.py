"""
migrate.py — טעינת נתונים מ-Excel ל-Supabase PostgreSQL

גליונות:
  ריכוז נתונים  — כותרות שורה 4, נתונים שורה 5+
  חברים         — כותרות שורה 5, נתונים שורה 6+
  הכנסות שכר   — כותרות שורה 8, נתונים שורה 9+
  סיסמאות       — כותרות שורה 1, נתונים שורה 3+

הרצה:
  python migrate.py
"""

import sys
import bcrypt
import psycopg2
from psycopg2.extras import execute_values
from openpyxl import load_workbook
from datetime import datetime

EXCEL_PATH   = 'masuot-data.xlsx'

# ערכים שנחשבים כ-NULL — כל ערך שמופיע כאן יהפוך None בDB
NULL_VALUES = {'--', '-', 'N/A', 'n/a', 'none', 'null', '0-', ''}


def get_connection():
    conn = psycopg2.connect(
        host="aws-1-ap-northeast-1.pooler.supabase.com",
        port=5432,
        database="postgres",
        user="postgres.jhkxyiiwtxtgqxovljkl",
        password="__Por@t2019!"
    )
    return conn

SCHEMA_SQL = """
DROP TABLE IF EXISTS salary_income CASCADE;
DROP TABLE IF EXISTS members CASCADE;
DROP TABLE IF EXISTS users CASCADE;
DROP TABLE IF EXISTS families CASCADE;
DROP TABLE IF EXISTS constants CASCADE;

CREATE TABLE families (
  budget_code         VARCHAR(20) PRIMARY KEY,
  family_name         VARCHAR(100),
  status              VARCHAR(20),
  budget_type         VARCHAR(20),
  budget_distribution DECIMAL(10,2) DEFAULT 0,
  special_help        DECIMAL(10,2) DEFAULT 0,
  periodic_grant      DECIMAL(10,2) DEFAULT 0,
  personal_bonus      DECIMAL(10,2) DEFAULT 0,
  women_work_benefit  DECIMAL(10,2) DEFAULT 0,
  travel              DECIMAL(10,2) DEFAULT 0,
  pension             DECIMAL(10,2) DEFAULT 0,
  survivors           DECIMAL(10,2) DEFAULT 0,
  old_age_pension     DECIMAL(10,2) DEFAULT 0,
  community_tax       DECIMAL(10,2) DEFAULT 0,
  property_tax        DECIMAL(10,2) DEFAULT 0,
  created_at TIMESTAMP DEFAULT NOW(),
  updated_at TIMESTAMP DEFAULT NOW()
);

CREATE TABLE members (
  id                  SERIAL PRIMARY KEY,
  member_code         VARCHAR(20) UNIQUE NOT NULL,
  budget_code         VARCHAR(20) NOT NULL
                      REFERENCES families(budget_code) ON DELETE CASCADE,
  first_name          VARCHAR(50),
  last_name           VARCHAR(50),
  gender              CHAR(1),
  birth_date          DATE,
  age                 INTEGER,
  status_code         INTEGER,
  status_name         VARCHAR(50),
  family_status       VARCHAR(20),
  education_group     VARCHAR(20),
  old_age_pension     DECIMAL(10,2) DEFAULT 0,
  pension_income      DECIMAL(10,2) DEFAULT 0,
  survivors_income    DECIMAL(10,2) DEFAULT 0,
  count_as_person     BOOLEAN DEFAULT TRUE,
  created_at          TIMESTAMP DEFAULT NOW()
);

CREATE INDEX idx_members_budget_code ON members(budget_code);

CREATE TABLE salary_income (
  id                    SERIAL PRIMARY KEY,
  budget_code           VARCHAR(20) NOT NULL
                        REFERENCES families(budget_code) ON DELETE CASCADE,
  member_code           VARCHAR(20),
  member_name           VARCHAR(100),
  gross_income          DECIMAL(10,2) DEFAULT 0,
  gross_income_expected DECIMAL(10,2) DEFAULT 0,
  work_percentage       DECIMAL(5,2)  DEFAULT 1,
  credit_points         DECIMAL(5,2)  DEFAULT 0,
  classification        VARCHAR(20),
  net_salary            DECIMAL(10,2) DEFAULT 0,
  net_salary_updated    DECIMAL(10,2) DEFAULT 0,
  training_fund         DECIMAL(10,2) DEFAULT 0,
  pension_addition      DECIMAL(10,2) DEFAULT 0,
  created_at            TIMESTAMP DEFAULT NOW()
);

CREATE INDEX idx_salary_budget_code ON salary_income(budget_code);

CREATE TABLE constants (
  id                       INTEGER PRIMARY KEY DEFAULT 1,
  CHECK (id = 1),
  price_daycare            DECIMAL(10,2) DEFAULT 3041,
  price_kindergarten       DECIMAL(10,2) DEFAULT 773,
  price_elementary         DECIMAL(10,2) DEFAULT 675,
  price_middle             DECIMAL(10,2) DEFAULT 1040,
  price_high               DECIMAL(10,2) DEFAULT 1229,
  edu_max_pct              DECIMAL(5,4)  DEFAULT 0.20,
  price_health_regular     DECIMAL(10,2) DEFAULT 193,
  price_health_young       DECIMAL(10,2) DEFAULT 116,
  price_health_middle      DECIMAL(10,2) DEFAULT 185,
  price_health_senior      DECIMAL(10,2) DEFAULT 825,
  health_max_pct           DECIMAL(5,4)  DEFAULT 0.075,
  balance_tax_threshold_1  DECIMAL(10,2) DEFAULT 10000,
  balance_tax_threshold_2  DECIMAL(10,2) DEFAULT 20000,
  balance_tax_threshold_3  DECIMAL(10,2) DEFAULT 99000,
  balance_tax_rate_1       DECIMAL(5,4)  DEFAULT 0.10,
  balance_tax_rate_2       DECIMAL(5,4)  DEFAULT 0.10,
  balance_tax_rate_3       DECIMAL(5,4)  DEFAULT 0.10,
  balance_tax_max          DECIMAL(10,2) DEFAULT 4000,
  tkh_threshold            DECIMAL(10,2) DEFAULT 4000,
  community_tax_per_member DECIMAL(10,2) DEFAULT 750,
  training_fund_threshold  DECIMAL(10,2) DEFAULT 17500,
  updated_at TIMESTAMP DEFAULT NOW()
);

CREATE TABLE users (
  id            SERIAL PRIMARY KEY,
  budget_code   VARCHAR(20) UNIQUE NOT NULL
                REFERENCES families(budget_code) ON DELETE CASCADE,
  description   VARCHAR(100),
  password_hash VARCHAR(255) NOT NULL,
  role          VARCHAR(20) DEFAULT 'member'
                CHECK (role IN ('member', 'admin')),
  last_login    TIMESTAMP,
  created_at    TIMESTAMP DEFAULT NOW()
);
"""


def create_schema(conn):
    print("Creating schema...")
    with conn.cursor() as cur:
        cur.execute(SCHEMA_SQL)
    conn.commit()
    print("  ✓ Schema created")


# ── פונקציות עזר ──────────────────────────────

def safe_decimal(value):
    """המר לספרה. None / ערכים ריקים → 0."""
    if value is None:
        return 0
    if isinstance(value, str) and value.strip() in NULL_VALUES:
        return 0
    try:
        return float(value)
    except (ValueError, TypeError):
        return 0

def safe_str(value):
    """המר לטקסט. None / ערכים ריקים / '--' וכד' → None."""
    if value is None:
        return None
    s = str(value).strip()
    if s in NULL_VALUES:
        return None
    return s or None

def safe_int(value):
    """המר למספר שלם. None / שגיאה → None."""
    if value is None:
        return None
    if isinstance(value, str) and value.strip() in NULL_VALUES:
        return None
    try:
        return int(float(value))
    except (ValueError, TypeError):
        return None

def safe_education_group(value):
    """
    קבוצת חינוך — רק ערכים תקינים עוברים.
    כל שאר הערכים (כולל '--', '-', ריק) → None.
    """
    VALID_GROUPS = {'פעוטון', 'גנים', 'יסודי', 'חטיבה', 'תיכון'}
    s = safe_str(value)
    if s in VALID_GROUPS:
        return s
    return None

def hash_password(password):
    hashed = bcrypt.hashpw(password.encode('utf-8'), bcrypt.gensalt())
    if isinstance(hashed, bytes):
        return hashed.decode('utf-8')
    return hashed


# ── 1. families ────────────────────────────────
def load_families(conn, wb):
    print("\nLoading families...")
    ws = wb['ריכוז נתונים']
    headers = [cell.value for cell in ws[4]]

    def col(name):
        return headers.index(name)

    rows = []
    skipped = 0

    for row_num in range(5, ws.max_row + 1):
        row = [ws.cell(row=row_num, column=i+1).value for i in range(len(headers))]
        budget_code = safe_str(row[col('קוד תקציב')])
        if not budget_code:
            skipped += 1
            continue

        rows.append((
            budget_code,
            safe_str(row[col('תאור תקציב')]),
            safe_str(row[col('מצב')]),
            safe_str(row[col('סוג תקציב')]),
            safe_decimal(row[col('חלוקת תקציב')]),
            safe_decimal(row[col('מיוחד-עזרה בבית')]),
            safe_decimal(row[col('מענק תקופתי')]),
            safe_decimal(row[col('תגמול אישי')]),
            safe_decimal(row[col('סל הטבות עבודת האישה')]),
            safe_decimal(row[col('נסיעות')]),
            safe_decimal(row[col('פנסיה')]),
            safe_decimal(row[col('שארים')]),
            safe_decimal(row[col('קצבת זקנה')]),
            safe_decimal(row[col('מס קהילה')]),
            safe_decimal(row[col('ארנונה')]),
        ))

    with conn.cursor() as cur:
        execute_values(cur, """
            INSERT INTO families (
                budget_code, family_name, status, budget_type,
                budget_distribution, special_help, periodic_grant,
                personal_bonus, women_work_benefit, travel,
                pension, survivors, old_age_pension,
                community_tax, property_tax
            ) VALUES %s
            ON CONFLICT (budget_code) DO NOTHING
        """, rows)

    conn.commit()
    print(f"  ✓ {len(rows)} families loaded ({skipped} skipped)")


# ── 2. members ─────────────────────────────────
def load_members(conn, wb):
    print("\nLoading members...")
    ws = wb['חברים']
    headers = [cell.value for cell in ws[5]]

    def col(name):
        try:
            return headers.index(name)
        except ValueError:
            return None

    with conn.cursor() as cur:
        cur.execute("SELECT budget_code FROM families")
        valid_codes = {row[0] for row in cur.fetchall()}

    rows = []
    skipped = 0

    for row_num in range(6, ws.max_row + 1):
        row = [ws.cell(row=row_num, column=i+1).value for i in range(len(headers))]

        member_code = safe_str(row[col('קוד חבר')]) if col('קוד חבר') is not None else None
        budget_code = safe_str(row[col('קוד תקציב')]) if col('קוד תקציב') is not None else None

        if not member_code or not budget_code or budget_code not in valid_codes:
            skipped += 1
            continue

        birth_date = row[col('תאריך לידה')] if col('תאריך לידה') is not None else None
        if isinstance(birth_date, datetime):
            birth_date = birth_date.date()
        elif birth_date:
            try:
                birth_date = datetime.strptime(str(birth_date), '%Y-%m-%d').date()
            except:
                birth_date = None

        # ← תיקון מרכזי: safe_education_group במקום safe_str
        edu_col = col('קבוצת חינוך')
        education_group = safe_education_group(row[edu_col]) if edu_col is not None else None

        rows.append((
            member_code, budget_code,
            safe_str(row[col('שם')]) if col('שם') is not None else None,
            safe_str(row[col('שם משפחה')]) if col('שם משפחה') is not None else None,
            safe_str(row[col('מין')]) if col('מין') is not None else None,
            birth_date,
            safe_int(row[col('גיל')]) if col('גיל') is not None else None,
            safe_int(row[col('קוד מעמד')]) if col('קוד מעמד') is not None else None,
            safe_str(row[col('מעמד')]) if col('מעמד') is not None else None,
            safe_str(row[col('מצב משפחתי')]) if col('מצב משפחתי') is not None else None,
            education_group,  # ← עכשיו תמיד NULL או ערך תקין בלבד
            safe_decimal(row[col('קצבת זיקנה בניכוי בריאות')]) if col('קצבת זיקנה בניכוי בריאות') is not None else 0,
            safe_decimal(row[col('פנסיה')]) if col('פנסיה') is not None else 0,
            safe_decimal(row[col('שארים')]) if col('שארים') is not None else 0,
        ))

    with conn.cursor() as cur:
        execute_values(cur, """
            INSERT INTO members (
                member_code, budget_code,
                first_name, last_name, gender, birth_date, age,
                status_code, status_name, family_status,
                education_group,
                old_age_pension, pension_income, survivors_income
            ) VALUES %s
            ON CONFLICT (member_code) DO NOTHING
        """, rows)

    conn.commit()
    print(f"  ✓ {len(rows)} members loaded ({skipped} skipped)")


# ── 3. salary_income ───────────────────────────
def load_salary_income(conn, wb):
    print("\nLoading salary income...")
    ws = wb['הכנסות שכר ']
    headers = [cell.value for cell in ws[8]]

    def col(name):
        try:
            return headers.index(name)
        except ValueError:
            return None

    with conn.cursor() as cur:
        cur.execute("SELECT budget_code FROM families")
        valid_codes = {row[0] for row in cur.fetchall()}

    rows = []
    skipped = 0

    for row_num in range(9, ws.max_row + 1):
        row = [ws.cell(row=row_num, column=i+1).value for i in range(len(headers))]

        budget_code = safe_str(row[col('קוד תקציב')]) if col('קוד תקציב') is not None else None
        if not budget_code or budget_code not in valid_codes:
            skipped += 1
            continue

        gross     = safe_decimal(row[col('הכנסה לחישוב שווי')]) if col('הכנסה לחישוב שווי') is not None else 0
        gross_exp = safe_decimal(row[col('הכנסה לאחר גידול')]) if col('הכנסה לאחר גידול') is not None else gross
        net       = safe_decimal(row[col('סכום נטו')]) if col('סכום נטו') is not None else 0
        # נטו לאחר גידול = שכר נטו כולל גידול שכר צפוי (העמודה הנכונה לסימולציה)
        net_upd   = safe_decimal(row[col('נטו לאחר גידול')]) if col('נטו לאחר גידול') is not None else net

        rows.append((
            budget_code,
            safe_str(row[col('קוד חבר')]) if col('קוד חבר') is not None else None,
            safe_str(row[col('שם')]) if col('שם') is not None else None,
            gross, gross_exp or gross,
            safe_decimal(row[col('אחוז משרה')]) if col('אחוז משרה') is not None else 1,
            safe_decimal(row[col('נקודות זיכוי')]) if col('נקודות זיכוי') is not None else 0,
            safe_str(row[col('סיווג')]) if col('סיווג') is not None else None,
            net, net_upd or net,
            safe_decimal(row[col('תוספת קרן השתלמות')]) if col('תוספת קרן השתלמות') is not None else 0,
            safe_decimal(row[col('הפרשות פנסיה')]) if col('הפרשות פנסיה') is not None else 0,
        ))

    with conn.cursor() as cur:
        execute_values(cur, """
            INSERT INTO salary_income (
                budget_code, member_code, member_name,
                gross_income, gross_income_expected,
                work_percentage, credit_points, classification,
                net_salary, net_salary_updated,
                training_fund, pension_addition
            ) VALUES %s
        """, rows)

    conn.commit()
    print(f"  ✓ {len(rows)} salary records loaded ({skipped} skipped)")


# ── 4. constants ───────────────────────────────
def load_constants(conn):
    print("\nLoading constants...")
    with conn.cursor() as cur:
        cur.execute("INSERT INTO constants (id) VALUES (1) ON CONFLICT (id) DO NOTHING")
    conn.commit()
    print("  ✓ Constants loaded with default values")


# ── 5. users ───────────────────────────────────
def load_users(conn, wb):
    print("\nLoading users from sheet 'סיסמאות'...")
    ws = wb['סיסמאות']

    with conn.cursor() as cur:
        cur.execute("SELECT budget_code FROM families")
        valid_codes = {row[0] for row in cur.fetchall()}

    rows = []
    skipped = 0

    for row_num in range(3, ws.max_row + 1):
        budget_code = safe_str(ws.cell(row=row_num, column=1).value)
        description = safe_str(ws.cell(row=row_num, column=2).value)
        password    = safe_str(ws.cell(row=row_num, column=3).value)

        if not budget_code or not password:
            skipped += 1
            continue

        if budget_code not in valid_codes:
            skipped += 1
            continue

        rows.append((budget_code, description, hash_password(password), 'member'))

        if len(rows) % 50 == 0:
            print(f"  ... {len(rows)} processed")

    with conn.cursor() as cur:
        execute_values(cur, """
            INSERT INTO users (budget_code, description, password_hash, role)
            VALUES %s
            ON CONFLICT (budget_code) DO NOTHING
        """, rows)

    conn.commit()
    print(f"  ✓ {len(rows)} users loaded ({skipped} skipped)")


# ── MAIN ───────────────────────────────────────
def main():
    print("=" * 50)
    print("Migration: Excel → Supabase PostgreSQL")
    print("=" * 50)

    print(f"\nOpening Excel: {EXCEL_PATH}")
    try:
        wb = load_workbook(EXCEL_PATH, data_only=True)
        print(f"  ✓ Sheets: {wb.sheetnames}")
    except FileNotFoundError:
        print(f"ERROR: File not found: {EXCEL_PATH}")
        sys.exit(1)

    print(f"\nConnecting to database...")
    try:
        conn = get_connection()
        print("  ✓ Connected")
    except Exception as e:
        print(f"ERROR: {e}")
        sys.exit(1)

    try:
        create_schema(conn)
        load_families(conn, wb)
        load_members(conn, wb)
        load_salary_income(conn, wb)
        load_constants(conn)
        load_users(conn, wb)

        print("\n" + "=" * 50)
        print("✓ Migration completed successfully!")
        print("=" * 50)

        with conn.cursor() as cur:
            cur.execute("SELECT COUNT(*) FROM families")
            print(f"  families:      {cur.fetchone()[0]}")
            cur.execute("SELECT COUNT(*) FROM members")
            print(f"  members:       {cur.fetchone()[0]}")
            cur.execute("SELECT COUNT(*) FROM salary_income")
            print(f"  salary_income: {cur.fetchone()[0]}")
            cur.execute("SELECT COUNT(*) FROM users")
            print(f"  users:         {cur.fetchone()[0]}")
            print(f"  constants:     1 row")

        # ── בדיקת sanity אחרי מיגרציה ──
        print("\n── Sanity checks ──")
        with conn.cursor() as cur:
            cur.execute("""
                SELECT COUNT(*) FROM members
                WHERE education_group IS NOT NULL
                AND education_group NOT IN ('פעוטון','גנים','יסודי','חטיבה','תיכון')
            """)
            bad_edu = cur.fetchone()[0]
            if bad_edu > 0:
                print(f"  ⚠️  {bad_edu} members with invalid education_group")
            else:
                print(f"  ✓ education_group values all valid")

            cur.execute("SELECT COUNT(*) FROM members WHERE education_group IS NULL")
            print(f"  ✓ {cur.fetchone()[0]} members with NULL education_group (adults/no group)")

            cur.execute("SELECT COUNT(*) FROM members WHERE education_group IS NOT NULL")
            print(f"  ✓ {cur.fetchone()[0]} members with education group (children)")

    except Exception as e:
        conn.rollback()
        print(f"\nERROR: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)
    finally:
        conn.close()


if __name__ == '__main__':
    main()